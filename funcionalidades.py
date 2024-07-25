
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import datetime as dt
import numpy as np
#("¯\_(ツ)_/¯")
greenColour = "\x1b[3;32m"#formato hexadecimal
yellowColour = '\033[1;33m'#formato octal
redColor= "\033[1;31m"
endColour = "\033[0m"#finaliza color
whiteColour = "\033[1;37m"
cianColour = "\033[1;36m"
blackColour = "\033[1;30m"
blueColour = "\033[1;34m"
purpleColour = "\033[1;35m"
fondoblanco = "\033[1;32;47m"
meses = {
    1 :'enero',
    2:'febrero',
    3:'marzo',
    4:'abril',
    5:'mayo',
    6:'junio',
    7:'julio',
    8:'agosto',
    9:'septiembre',
    10:'octubre',
    11:'noviembre',
    12:'diciembre'
}


def saveInExcel(archivo,data,nameSheet, heading):
    """esta funcion recive una serie de parametros donde se obtiene el archivo la data el nombre de la hoja y el encabezado, ya ejecuta la libreria openpyxl y load workbook para escribir dentro del archivo de excel
    
    """
    
    try:
        print(heading)

        wb = load_workbook(archivo)
        ws = wb[nameSheet]
        headings = list(data.keys()) 
        if heading == 's':
            ws.append(headings)
        lista = list(data.values())
        ws.append(lista)
            
        
        wb.save(archivo)
        
    except Exception as e :
        print(e)
def menssages(color,msg):
    print(f"{color} {msg} {endColour}")

def resetasCosto(DFs,*args, **kwargs):

    """ esta funcion sirve para sumar los ingredientes y sumarle costos al producto arroja el costo total del producto se debe introducir la data con un archivo excel 1ero introduce las unidades que arroja la receta 2do las unidades por paquete 3ero costo del empaque y el costo de la publicidad """
    data = pd.DataFrame(DFs)
    data.fillna(0, inplace= True)
    data.columns = ['Ingredientes', 'cantidad', 'costo', 'valor']
    total = data['valor'].sum()
    cantidad = data['cantidad'].sum()
    gramaje = 0 if args[4] == "" else int(args[4])  
    
    produccionTotalArrojadas = cantidad / gramaje

    
    unidadesArrojadasXreceta = int(args[0])
    valorUnitario = total / unidadesArrojadasXreceta
    unidades = int(args[1])
    empaque = args[2]
    publicidad = args[3]

    empacadoValor = valorUnitario * unidades + empaque + publicidad
    veintePorciento = empacadoValor * 20 /100
    totalProducto = empacadoValor + veintePorciento
    return totalProducto, produccionTotalArrojadas


def crear_grafico():
    data = input(f"{greenColour} Elige la data: [1] registros [2]ventas {endColour}")
    if data == '1':
        try:
            datos = pd.read_excel(".\\ResetasPrueba.xlsx",sheet_name='registros', engine='openpyxl')
        
            tipo_De = input(f"{whiteColour} tipo De grafico [0] de torta [1] lineal [2] barras {endColour}")
            
        except Exception as E:
            print(f"{redColor} {E} {endColour}")
        if tipo_De == '0':
            df = pd.DataFrame(datos)
            dfNan = df.dropna().copy()
           
            dfNanvlues = dfNan.drop(['fecha','Producto','porcentaje_ganancia','cantidad','costoProductofinal','cuentaTotal'],axis=1)
            headers = dfNanvlues.columns.values
            values = dfNanvlues.sum().values
            
            print(headers)
            explode = [0.1, 0.00]
            colors = ['firebrick','green']
            fig, ax = plt.subplots()
            plt.pie(values,labels=headers,labeldistance=1,explode=explode, autopct='%.2f%%', pctdistance=0.8, colors=colors,startangle=100)
            plt.show()
    elif data == '2':
        try:
            datos = pd.read_excel(".\\ResetasPrueba.xlsx",sheet_name='ventas', engine='openpyxl')
        except Exception as E:
            print(f"{redColor} {E} {endColour}")        

        df_ventas= pd.DataFrame(datos).copy()
        df_ventas['venta total'] = (df_ventas['cantidad'] * df_ventas['precio'])
            # df_ventas = df_ventas.drop(df_ventas[(df_ventas['producto']=='Ninguno') | (df_ventas['producto']=='ninguno')].index)
        
        df_ventas['date'] = df_ventas['date'].apply(lambda x : pd.to_datetime(x,format="%d/%m/%Y"))

        tipo_De = input(f"{whiteColour} tipo De grafico [0] de torta [1]lineal  [2] barras {endColour}")
        if tipo_De =='0':
            #months_sales = df_ventas.resample('M',on='date').sum() mes
            #months_sales = df_ventas.resample('D',on='date').sum() dia
            #months_sales = df_ventas.resample('w',on='date').sum() semana
            df_ventas = df_ventas.drop(df_ventas[(df_ventas['producto']=='Ninguno') | (df_ventas['producto']=='ninguno')].index)
            DfVentaTotal = df_ventas['venta total']
            Dfcantidad = df_ventas['cantidad']
            Dffecha = df_ventas.iloc[:,0]
            
            newDf = pd.concat([Dffecha,DfVentaTotal ,Dfcantidad ],axis=1)
            months_sales = newDf.resample('w',on='date').sum()
            print(months_sales)

        elif tipo_De == '1':
           
            mes = int(input(f"{whiteColour}que mes deseas graficar:  {endColour}"))
            selMes = df_ventas[df_ventas['date'].dt.month == mes]

            cantidad = list(df_ventas['cantidad'])
            
            fecha = list(selMes['date'])
            fig, ax = plt.subplots()
            
            plt.plot(fecha,cantidad,color="g", linestyle="solid", marker="o", label="ventas")
            plt.xlabel("Fechas")
            plt.xticks(rotation=25)
            plt.ylabel("Cantidad")
            plt.title("Ventas mes de {}".format(meses[mes]))
            plt.legend()
            plt.grid(True)
            plt.xticks(fecha)
            plt.show()
            
            
    else:
        print('error')
def insertar_venta():
    salir = False
    print(salir)
    while  salir != True:
        try:
            print()
        
            
            wb = load_workbook(".\\ResetasPrueba.xlsx")
            ws = wb['ventas']
            fecha = input(f"{whiteColour} introduce la feche si deseas la{yellowColour} fecha {endColour} de hoy marca 's': {endColour}\n") 
            fechanow = dt.date.today() if fecha == 's' else pd.to_datetime(fecha, format="%d/%m/%Y",dayfirst=True).date()
            producto1 = input(f"{whiteColour} introduce el{blueColour} nombre del producto {endColour} presiona enter Para 'Ninguno': {endColour}\n")
            producto = 'Ninguno' if producto1 == '' else producto1
            cantidad1 = input(f"{whiteColour} introduce la {greenColour} cantidad {endColour} presiona enter Para '0': {endColour}\n")
            cantidad = 0 if cantidad1 == '' else float(cantidad1)
            precio1 = input(f"{whiteColour} introduce el {purpleColour}precio {endColour} presiona enter Para '0': {endColour}\n")
            precio = 0 if precio1 == '' else float(precio1)
            metodo_de_pago = input(f"{whiteColour} introduce  {cianColour}metodo de pago {endColour} 1 para:'fiado' o 2 para:'contado' presiona para ninguno : {endColour}\n")
            tipo_de_pago = 'fiado' if metodo_de_pago == '1' else 'Ninguno' if metodo_de_pago == '' else 'contado'
            Negocio_consumidor = input(f"{whiteColour} introduce 'c' si es un {redColor} cliente si es un negocio {endColour} su nombre  presiona enter Para 'Ninguno': {endColour}\n")
            Negocio_o_consumidor = 'cliente' if Negocio_consumidor == 'c' else 'Ninguno' if Negocio_consumidor == '' else Negocio_consumidor

            df = {
                'fecha':fechanow,
                'producto':producto,
                'cantidad':cantidad,
                'precio':precio,
                'metodo_de_pago':tipo_de_pago,
                'Negocio_consumidor': Negocio_o_consumidor
            }
            
            
            lista = list(df.values())
            print(f"{yellowColour} informacio ingresada {endColour} {greenColour}{lista}{endColour} ")
            ws.append(lista)
                
            
            wb.save(".\\ResetasPrueba.xlsx")
            menssages(redColor,"Deseas continuar insertando marca enter para SI y N para salir")
            seguir = input(": ")
            if seguir:
                return print(f'{greenColour}venta registrada{endColour}')
            else:
                continue
                
        
        except Exception as e :
            print(f"{redColor}{e.__class__} , {e} {endColour}")

def gestion_datos():
        
    
    try:
        dataset = pd.read_excel(".\\ResetasPrueba.xlsx",sheet_name='ventas', engine='openpyxl')
    except Exception as E:
        print(f"{redColor} {E} {endColour}")
    
    salir = False
    dataframe = pd.DataFrame(dataset)
    copy = dataframe.dropna().copy()
    copy['venta total'] = (copy['cantidad'] * copy['precio'])
    copy['date'] = copy['date'].apply(lambda x : pd.to_datetime(x,format="%d/%m/%Y"))
    while salir != True:
        
        opcion = input(f"{whiteColour}[0]mostrar datos {blueColour}[1]Producto Mayor venta {endColour}  {purpleColour}[2]Con menor Venta{endColour}\n\n {yellowColour}[3]dia con menor venta {endColour}{endColour}{greenColour} [4] dia con mayor venta {endColour} \n\n {blackColour}[5]tienda o cliente con mayor venta {endColour} {purpleColour}[6]tienda o cliente con menor compra {endColour}")
        

        if opcion == '0':
            
            cantidad_lineas = input(f"{yellowColour}introduce cuantas lineas deseas ver presiona 'enter' para ver todas: {endColour}")
            lineas = None if cantidad_lineas == '' else int(cantidad_lineas)
            print(f"{copy.iloc[0:lineas,[0, 2,-1]]}")
        elif opcion == '1':
            copy = copy.drop(copy[(copy['producto']=='Ninguno') | (copy['producto']=='ninguno')].index)
            
            maximo = copy.groupby('producto')['venta total'].sum()
            cantidadMax = copy.groupby('producto')['cantidad'].sum()
            return print(f"{greenColour} El producto con mayor venta es:{fondoblanco} {maximo.idxmax()} con un Total de: {maximo.max()}$ Generado{endColour}\n\n y una cantidad total de {fondoblanco} {cantidadMax.max()} Paquetes{endColour} {endColour}")

        elif opcion == '2':
            copy = copy.drop(copy[(copy['producto']=='Ninguno') | (copy['producto']=='ninguno')].index)
            minimo = copy.groupby('producto')['venta total'].sum()
            cantidadMin = copy.groupby('producto')['cantidad'].sum()
            print(f"{greenColour} El producto con menor venta es:{fondoblanco} {minimo.idxmin()} con un Total de: {minimo.min()}$ Generado{endColour}\n\n y una cantidad total de {fondoblanco} {cantidadMin.min()} Paquetes{endColour} {endColour}")
        elif opcion == '3':
            dfcantidad = copy['cantidad']
            dffecha = copy['date']
            dfproducto = copy['producto']
            dfTotalVenta = copy['cantidad'] * copy['precio']
            newdf = pd.concat([dffecha,dfcantidad, dfproducto],axis=1)
            newdf['totalventa'] = dfTotalVenta
            
            newdf= newdf.drop(newdf[(newdf['producto']=='Ninguno') & (newdf['cantidad']== 0)].index )
            selmes = input("introdusca el mes con el cual trabajar: ")
            #selecion de mes 
            mes = newdf[newdf['date'].dt.month == int(selmes)].copy()
            day_sales = mes.resample('D',on='date').sum().sort_values(by='totalventa').copy()
            day_sales = day_sales.drop(day_sales[(day_sales['totalventa'] == 0)].index)
            
            min_say_day = day_sales[day_sales['totalventa'] == day_sales['totalventa'].min()].iloc[0]

            """retorno de la funcion"""
            print(f"\n {yellowColour}el dia con menor venta es:{endColour} {redColor} {pd.to_datetime(min_say_day.name, format="%d/%m/%Y",dayfirst=True).date()}{endColour} con una venta total de: {greenColour}{min_say_day.array[2]}${endColour} el producto para este dia es:{fondoblanco} {min_say_day.array[1][0:10]} {endColour}y su cantidad es de:{blueColour} {min_say_day.array[0]} {endColour}unidades|paquetes\n ")

        elif opcion == '4':
            cantidad = copy['cantidad']
            dfecha = copy['date']
            dproducto = copy['producto']
            dTotalVenta = copy['cantidad'] * copy['precio']
            newd = pd.concat([dfecha,cantidad, dproducto],axis=1)
            newd['totalventa'] = dTotalVenta
            
            newd= newd.drop(newd[(newd['producto']=='Ninguno') & (newd['cantidad']== 0)].index )
            selme = input("introdusca el mes con el cual trabajar: ")
            mes = newd[newd['date'].dt.month == int(selme)].copy()
            day_sales = mes.resample('D',on='date').sum().sort_values(by='totalventa').copy()
            day_sales = day_sales.drop(day_sales[(day_sales['totalventa'] == 0)].index)
            
            max_say_day = day_sales[day_sales['totalventa'] == day_sales['totalventa'].max()].iloc[0]

            print(f"\n {greenColour}el dia con mayor venta es:{endColour} {whiteColour} {pd.to_datetime(max_say_day.name, format="%d/%m/%Y",dayfirst=True).date()}{endColour} con una venta total de: {greenColour}{max_say_day.array[2]}${endColour} el producto para este dia es:{fondoblanco} {max_say_day.array[1][0:10]} {endColour}y su cantidad es de:{blueColour} {max_say_day.array[0]} {endColour}unidades|paquetes\n ")

        elif opcion == '5':
            data = input(f"deseas saber el cliente o la tienda con mayor venta marca 'C' para cliente o 'T'para tienda ").lower()
            
            if data == 'c':
                copy['date'] = copy['date'].apply(lambda x : pd.to_datetime(x,format="%d/%m/%Y"))
                seleccionmes = input(f"{whiteColour}mes para trabajar con la data:  {endColour}")
                mes = copy[copy['date'].dt.month == int(seleccionmes)].copy()
                mes = mes.drop(mes[(mes['venta total'] == 0)].index)
               
                df_resume = pd.pivot_table(mes,
                    index=['date','Negocio o consumidor'],
                    fill_value=0,
                    columns='producto',
                    values=['venta total','cantidad'],
                    aggfunc='sum'
                )
              
                tienda_menor_venta = df_resume.idxmin()
                print(df_resume)
                print(tienda_menor_venta.min())
                # week = df_resume.resample('D', level=0).sum()
                
                
            # newdf = newdf.drop('producto',axis=1)
            # print(newdf.columns)
            # day_sales = newdf.resample('D',on='date').sum().sort_values(by='totalventa').copy()
            
            # print(day_sales.columns)
            # day_sales = day_sales.drop(day_sales[(day_sales['totalventa'] == 0)].index)
            # # print(day_sales.columns)
            # fecha = day_sales.head(1)
            # # fecha2 = fecha.axes[0][0]
            # mes = input("mes")
            # selMes =  day_sales.columns           
            #  # selMes = day_sales[day_sales['date'].dt.month == mes]
            # print(selMes)
            # selMes.groupby(['date','cantidad', 'producto','totalventa']).agg('min')
           
            # day_sales = day_sales.min(axis=1)
            # print(day_sales)
            # print(f" \n {yellowColour}el dia con menor venta es:{endColour} {fondoblanco}{fecha.axes[0][0]}{endColour}   con una cantidad de: {day_sales.array[0]} venta total de: {day_sales.array[1]} el producto: {day_sales.array[2]} ")
          


        else:
            return print(f'{yellowColour}saliendo.. {endColour}')
        

def inventario():
    salir = False
    while salir != True:
        print(f"{yellowColour}elija una de las siguientes opciones:\n {endColour}")
        manage = input(f"{whiteColour}[0] crear costo x materia prima{endColour} {blueColour} [1] inventario de materias prima {endColour} {blackColour} [2] stock de productos {endColour} {purpleColour}[3]produccion diaria {endColour}")
        wb = load_workbook(".\\ResetasPrueba.xlsx")
        ws = wb['inventario']
        if manage == '0':
            print(f"{yellowColour}[1]mostrar data [2]costear Mprima{endColour}")
            opcion = input(f"{whiteColour}Elige la opcion:  {endColour}")
            if opcion == '1':
              for row in ws.iter_rows():
                  print([cell.value for cell in row])  
            elif opcion == '2':
                name = input(f"{whiteColour}nombre: {endColour}")
                fecha_nueva = None
                for row in ws.iter_rows():
            # Verificar si el nombre del campo es igual al valor buscado
                    
                    if row[1].value == name:
                       fecha_actual = row[0].value
                       if fecha_nueva is None or fecha_actual > fecha_nueva:
                           fecha_nueva = fecha_actual
                           datos = [cell.value for cell in row]
                # Devolver todos los campos de la fila
                       
                # datos.append(['fecha','nombre', 'cantidad','kilos|gramos','precio'])
                dfCeldas = pd.DataFrame(datos,index=['fecha','nombre', 'cantidad','kilos|gramos','precio']).transpose()
                gramos = dfCeldas['kilos|gramos'].values
                valorgramos = gramos[0].split(' ')
                kilo = valorgramos[0] if len(valorgramos[0]) > 1 else valorgramos[0] + '000'
                if len(kilo) == 2:
                    precio_gramo = dfCeldas['precio'] / float(kilo) / 1000
                else:
                    precio_gramo = dfCeldas['precio'] / float(kilo)
                print(f"\n{blueColour} el precio por gramo para {dfCeldas['nombre'].values[0]}es: {precio_gramo.array[0]}{endColour}\n\n y la informacion recogida del inventario es la siguiente")
                print(dfCeldas)
                dts = [dfCeldas['nombre'].values[0],precio_gramo.array[0]]
                nws =wb['costo_por_materia']
                nws.append(dts)
                wb.save(".\\ResetasPrueba.xlsx")
        elif manage == '1':
            try:
                Pfecha = input(f"{whiteColour} fecha ingreso o presiona enter para ingresar la de hoy:  {endColour}")
                fecha = pd.to_datetime(Pfecha, format="%d/%m/%Y",dayfirst=True).date() if len(Pfecha) > 5 else dt.date.today()
                nombre = input(f"{whiteColour}nombre de la materia: {endColour}")
                cantidad =int(input(f"{whiteColour} cantidad:  {endColour}"))
                kilosOgramos = input(f"{whiteColour}cuantos kilos o gramos introduce solo el numero {endColour}")
                kgogr = input(f"{whiteColour} 'k' para kilos 'g' para gramos {endColour}")
                kg_gr = kilosOgramos + ' kilos' if kgogr == 'k' else kilosOgramos +' gramos'
                precio = float(input(f"{whiteColour}Precio:  {endColour}"))
                in_stock = float(input(f"{whiteColour} cantidad en stock: {endColour}"))
                df= {
                    'fecha':fecha,
                    'nombre':nombre,
                    'cantidad':cantidad,
                    'kg_gr':kg_gr,
                    'precio':precio,
                    'en_stock': in_stock
                }
                data = list(df.values())
                ws.append(data)
                wb.save(".\\ResetasPrueba.xlsx")
                print(f"{greenColour}materia prima registrada con exito{endColour}")
            except Exception as e:
                print(f"{redColor}{e.__class__} , {e} {endColour}")
        else:
            return print(f"{yellowColour}Saliendo... {endColour}")